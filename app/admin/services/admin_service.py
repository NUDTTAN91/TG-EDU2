from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from app.models.user import User
from app.models.course import Course
from app.models.submission import Submission
from app.models.assignment import Assignment
from app.models.school import School
from app.models.school_class import Class, class_students
from app.utils.security import get_password_hash


async def create_user(db: AsyncSession, username: str, full_name: str, role: str, password: str = "123456", school_id: int = None, class_id: int = None):
    existing = await db.execute(select(User).where(User.username == username))
    if existing.scalar_one_or_none():
        return None, "用户名已存在"
    user = User(
        username=username,
        full_name=full_name,
        password_hash=get_password_hash(password),
        role=role,
        school_id=school_id,
        must_change_password=(role != "admin"),
    )
    db.add(user)
    await db.flush()
    if class_id and role == "student":
        cls = await db.execute(select(Class).where(Class.id == class_id))
        cls_obj = cls.scalar_one_or_none()
        if cls_obj:
            await db.execute(
                class_students.insert().values(class_id=class_id, student_id=user.id)
            )
    await db.commit()
    await db.refresh(user)
    return user, None


async def get_users(db: AsyncSession):
    result = await db.execute(select(User).options(selectinload(User.classes)))
    return result.scalars().unique().all()


async def get_user_by_id(db: AsyncSession, user_id: int):
    result = await db.execute(select(User).where(User.id == user_id))
    return result.scalar_one_or_none()


async def toggle_user_status(db: AsyncSession, user_id: int, operator_id: int):
    if user_id == operator_id:
        return None, "不能禁用自己的账号"
    user = await get_user_by_id(db, user_id)
    if not user:
        return None, "用户不存在"
    user.is_active = not user.is_active
    db.add(user)
    await db.commit()
    return user, None


async def import_students(db: AsyncSession, students_data: list):
    count = 0
    for s in students_data:
        existing = await db.execute(select(User).where(User.username == s.username))
        if existing.scalar_one_or_none():
            continue
        student = User(
            username=s.username,
            full_name=s.full_name,
            password_hash=get_password_hash(s.password),
            role="student",
            must_change_password=True,
        )
        db.add(student)
        count += 1
    await db.commit()
    return count


async def get_stats(db: AsyncSession):
    total_users = (await db.execute(select(func.count(User.id)))).scalar()
    total_teachers = (await db.execute(select(func.count(User.id)).where(User.role == "teacher"))).scalar()
    total_students = (await db.execute(select(func.count(User.id)).where(User.role == "student"))).scalar()
    total_courses = (await db.execute(select(func.count(Course.id)))).scalar()
    total_submissions = (await db.execute(select(func.count(Submission.id)))).scalar()
    total_assignments = (await db.execute(select(func.count(Assignment.id)))).scalar()
    total_schools = (await db.execute(select(func.count(School.id)))).scalar()
    return {
        "total_users": total_users,
        "total_teachers": total_teachers,
        "total_students": total_students,
        "total_courses": total_courses,
        "total_submissions": total_submissions,
        "total_assignments": total_assignments,
        "total_schools": total_schools,
    }


# ============================================================================
# 批量导入用户：xlsx 模板生成 + 解析
# ============================================================================

# 模板中英文表头同义词（下标以小写匹配）
_HEADER_ALIASES = {
    "role": {"role", "角色"},
    "username": {"username", "account", "账号", "学号", "工号", "学号/工号", "账号（学号/工号）"},
    "full_name": {"full_name", "name", "姓名"},
    "class_name": {"class", "class_name", "班级", "班级名", "班级（学生填）"},
    "school_name": {"school", "school_name", "学校", "院校", "学校名", "院校名"},
}

_ROLE_MAP = {
    "student": "student", "学生": "student",
    "teacher": "teacher", "教师": "teacher", "老师": "teacher",
}

# 批量导入用户统一使用默认密码，首登强制改密
DEFAULT_IMPORT_PASSWORD = "123456"


def _norm_header(cell_value) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip().lower().replace(" ", "")


def _map_header_row(header_row) -> dict:
    """将表头行映射为 {字段名: 列序号}。列序号从 0 开始。"""
    result = {}
    for idx, cell in enumerate(header_row):
        raw = _norm_header(cell)
        if not raw:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if raw in {_norm_header(a) for a in aliases}:
                if field not in result:
                    result[field] = idx
                break
    return result


def build_user_import_template() -> bytes:
    """生成模板 xlsx，返回字节。仅在路由以 StreamingResponse 返回。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入"

    # 顶部说明（合并单元格）
    ws.merge_cells("A1:E1")
    hint = ws["A1"]
    hint.value = (
        "填写说明：角色可填「学生」或「教师」；学生必填班级，教师班级列留空；学校列必填；"
        "学校/班级如不存在将自动创建。初始密码统一为 123456，首次登录将强制修改。导入时请删除本行下方的示例。"
    )
    hint.font = Font(bold=False, color="666666", size=10)
    hint.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 42

    # 表头
    headers = ["角色", "账号（学号/工号）", "姓名", "班级（学生填）", "学校"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D1D1D")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 示例行
    examples = [
        ["学生", "stu001", "张三", "计算机2101", "示例学院"],
        ["学生", "stu002", "李四", "计算机2101", "示例学院"],
        ["教师", "tea001", "王老师", "", "示例学院"],
    ]
    for r, row in enumerate(examples, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # 列宽
    widths = [10, 22, 14, 20, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_users_from_xlsx(db: AsyncSession, file_bytes: bytes) -> dict:
    """解析 xlsx 并批量入库。返回创建统计与 skipped 报告。

    学校名从行内读取；学校/班级不存在时自动创建。
    失败行均以 skipped 记录并继续，不因一行错误全回滚。
    文件本身无法解析时 raise ValueError。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("服务器未安装 openpyxl，无法解析 xlsx")

    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"无法解析 xlsx 文件：{e}")

    ws = wb.active
    if ws is None:
        raise ValueError("xlsx 文件中没有可用的 sheet")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("xlsx 文件为空")

    # 自动定位表头：从前 5 行里找包含必需字段的第一行（兼容顶部说明行存在）
    header_row_idx = None
    header_map = {}
    for i, row in enumerate(rows[:5]):
        m = _map_header_row(row)
        if {"username", "full_name", "role", "school_name"}.issubset(m.keys()):
            header_row_idx = i
            header_map = m
            break

    if header_row_idx is None:
        raise ValueError("未找到合法表头行（需含「角色」「账号」「姓名」「学校」列）")

    data_rows = rows[header_row_idx + 1:]

    # 预加载全库学校 name→id、班级 (school_id, name)→id、已存在用户名
    school_rows = (await db.execute(select(School.id, School.name))).all()
    school_name_to_id = {name: sid for sid, name in school_rows}

    class_rows = (await db.execute(select(Class.id, Class.name, Class.school_id))).all()
    class_key_to_id = {(sid, name): cid for cid, name, sid in class_rows}

    existing_usernames = set()
    for row in (await db.execute(select(User.username))).all():
        existing_usernames.add(row[0])

    created_students = 0
    created_teachers = 0
    created_schools = 0
    created_classes = 0
    skipped = []
    seen_in_batch = set()

    for offset, row in enumerate(data_rows):
        row_no = header_row_idx + 2 + offset

        def _cell(field):
            idx = header_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            if val is None:
                return ""
            return str(val).strip()

        role_raw = _cell("role")
        username = _cell("username")
        full_name = _cell("full_name")
        class_name = _cell("class_name")
        school_name = _cell("school_name")

        # 全空行跳过（不计失败）
        if not any([username, full_name, role_raw, class_name, school_name]):
            continue

        if not username:
            skipped.append({"row": row_no, "username": "", "reason": "账号为空"})
            continue
        if not full_name:
            skipped.append({"row": row_no, "username": username, "reason": "姓名为空"})
            continue

        role = _ROLE_MAP.get(role_raw.lower(), None)
        if not role:
            skipped.append({"row": row_no, "username": username, "reason": f"角色非法：{role_raw}"})
            continue

        if not school_name:
            skipped.append({"row": row_no, "username": username, "reason": "学校名为空"})
            continue

        if username in existing_usernames or username in seen_in_batch:
            skipped.append({"row": row_no, "username": username, "reason": "用户名已存在"})
            continue

        # 学校：没就建
        sid = school_name_to_id.get(school_name)
        if sid is None:
            new_school = School(name=school_name)
            db.add(new_school)
            await db.flush()
            sid = new_school.id
            school_name_to_id[school_name] = sid
            created_schools += 1

        # 学生：班级必填；不存在自动建班
        class_id = None
        if role == "student":
            if not class_name:
                skipped.append({"row": row_no, "username": username, "reason": "学生需填班级名"})
                continue
            key = (sid, class_name)
            class_id = class_key_to_id.get(key)
            if class_id is None:
                new_class = Class(name=class_name, school_id=sid)
                db.add(new_class)
                await db.flush()
                class_id = new_class.id
                class_key_to_id[key] = class_id
                created_classes += 1

        # 密码统一使用默认值，首次登录强制改密
        pwd_to_use = DEFAULT_IMPORT_PASSWORD

        user = User(
            username=username,
            full_name=full_name,
            password_hash=get_password_hash(pwd_to_use),
            role=role,
            school_id=sid,
            must_change_password=True,
        )
        db.add(user)
        await db.flush()  # 拿 user.id 用于学生入班

        if role == "student" and class_id is not None:
            await db.execute(
                class_students.insert().values(class_id=class_id, student_id=user.id)
            )
            created_students += 1
        else:
            created_teachers += 1

        existing_usernames.add(username)
        seen_in_batch.add(username)

    await db.commit()

    return {
        "created_students": created_students,
        "created_teachers": created_teachers,
        "created_schools": created_schools,
        "created_classes": created_classes,
        "skipped": skipped,
    }
